#%%
import cv2
import numpy as np
from openpyxl import Workbook
import datetime
starttime = datetime.datetime.now()

#%%
import glob
import os
from os import walk

# 在記憶體中建立一個workbook物件，而且會至少建立一個 worksheet
wb = Workbook()
#獲取當前活躍的worksheet,預設就是第一個worksheet
ws = wb.active
# 設定 Exl 的欄位名稱
ws["B1"]="770nm_L"; ws["C1"]="770nm_R"; ws["D1"]="850nm_L"; ws["E1"]="850nm_R"; ws["F1"]="940nm_L"; ws["G1"]="940nm_R"
names = ["770L01.png","770L02.png","770R01.png","770R02.png","850L01.png","850L02.png","850R01.png","850R02.png","940L01.png","940L02.png","940R01.png","940R02.png"]
col = 2
r = 2

for sub in os.listdir(r"demo\data"):
    path = r"demo\data"
    save_path = r"demo\wrist"# a path for saving image

    path = os.path.join(path, sub)
    save_path = os.path.join(save_path, sub)
    if not os.path.isfile(save_path):
        os.makedirs(save_path)
    # print(path)

    cv_img = []
    i = 0
    a = 0
    for img in os.listdir(path):
        if os.path.join(path, img).endswith(".png"):
            img = cv2.imread(os.path.join(path, img))
            cv_img.append(img)
                                      
        
            #Do histogram equalization
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) #turn RGB into GRAY 
            hist,bins = np.histogram(gray.flatten(),256,[0,256])
            cdf = hist.cumsum()
            cdf_normalized = cdf * hist.max()/ cdf.max()
            cdf_m = np.ma.masked_equal(cdf,0)# 除去直方圖中的0值
            cdf_m = (cdf_m - cdf_m.min())*255/(cdf_m.max()-cdf_m.min())
            cdf = np.ma.filled(cdf_m,0).astype('uint8')# 將掩模處理掉的元素補為0
            img2 = cdf[gray.astype(np.uint8)]
        
            # blur_gray = cv2.GaussianBlur(img2, (101, 101), 0) # Gaussian filter, the kernel must be an odd number
            ret,thresh1 = cv2.threshold(img2,200,255,cv2.THRESH_BINARY)
        
            _, contours, hierarchy = cv2.findContours(thresh1,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
        
            try: hierarchy = hierarchy[0]
            except: hierarchy = []
        
            height, width = thresh1.shape
            min_x, min_y = width, height
            max_x = max_y = 0
        
            # computes the bounding box for the contour, and draws it on the frame,
            for contour, hier in zip(contours, hierarchy):
                (x,y,w,h) = cv2.boundingRect(contour)
                min_x, max_x = min(x, min_x), max(x+w, max_x)
                min_y, max_y = min(y, min_y), max(y+h, max_y)
        
        
            if max_x - min_x > 0 and max_y - min_y > 0:
                cv2.rectangle(img, (int(min_x*1.1), int(min_y*1.0)), (int(max_x*0.95), int(max_y*0.9)), (255, 0, 0), 2) # 畫出適當ROI
            
            x_range = int(max_x*0.95) - int(min_x*1.1)
            if int(max_y*0.9) - (int(min_y) + x_range) < abs(int(min_x*1.1) - int(max_x*0.95))/5:
                add = int(max_y*0.9) - int(min_y) - abs(int(min_x*1.1) - int(max_x*0.95))/3
                rect =img2 [(int(min_y) + int(add)):int(max_y*0.9), int(min_x*1.1):int(max_x*0.95)]  
                
            else:
                rect =img2 [(int(min_y) + x_range):int(max_y*0.9), int(min_x*1.1):int(max_x*0.95)]  
            
            print ("The Intensity of",names[a] + " is "  "%.2f" %rect.mean())            

            
            cv2.imwrite(os.path.join(save_path, "{}".format(names[a])),rect)
            a += 1
            if a == 12:
                a = 0
        
            if col <= 7 :        
                ws.cell (row = r, column = col).value = rect.mean()
                col += 1
            else :
                col = 2
                r += 1        
                ws.cell (row = r, column = col).value = rect.mean()
                col += 1

#%% show images
cv2.namedWindow("Thresh1", 0)
cv2.resizeWindow("Thresh1", 640, 480)
cv2.imshow("Thresh1", thresh1)

cv2.namedWindow("Image", 0)
cv2.resizeWindow("Image", 640, 480)
cv2.imshow("Image", img)

#cv2.imwrite(filename, img)
cv2.waitKey(0)
cv2.destroyAllWindows()

